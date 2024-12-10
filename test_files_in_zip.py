import csv
import io

import pytest
import os
import script_os
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook

zip_file_path = os.path.join(script_os.RESOURCES_PATH, 'example.zip')


@pytest.fixture(scope='session')
def add_file_zip():
    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for file in os.listdir(script_os.DATA_PATH):
            file_path = os.path.join(script_os.DATA_PATH, file)
            zip_file.write(file_path, file)
    yield
    if os.path.exists(zip_file_path):
        os.remove(zip_file_path)


def test_pdf_in_zip(add_file_zip):
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        file_list = zip_ref.namelist()
        for file_name in file_list:
            if '.pdf' in file_name:
                pdf_name = file_name
                continue
        with zip_ref.open(pdf_name) as pdf_file:
            pdf_file = PdfReader(pdf_file)
            text = pdf_file.pages[0].extract_text()

        assert 'This document and PDF form have been created with OpenOffice' in text, \
            "Содержимое файла PDF не соответствует ожидаемому"


def test_xlsx_in_zip(add_file_zip):
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        file_list = zip_ref.namelist()
        for file_name in file_list:
            if '.xlsx' in file_name:
                xlsx_name = file_name
                continue
        with zip_ref.open(xlsx_name) as xlsx_file:
            xlsx_file = load_workbook(xlsx_file).active
            text = xlsx_file.cell(row=4, column=2).value

        assert 'Project Management Data' == text, 'Заголовок таблицы не соответствует ожидаемому'


def test_csv_in_zip(add_file_zip):
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        file_list = zip_ref.namelist()
        for file_name in file_list:
            if '.csv' in file_name:
                csv_name = file_name
                continue
        with zip_ref.open(csv_name) as csv_file:
            with io.TextIOWrapper(csv_file, encoding='utf-8') as text_file:
                text_file = csv.reader(text_file)
                headers = next(text_file)

        assert ['Name', 'Job Title', 'Address', 'State', 'City'] == headers, \
            'Список столбцов не соответствует ожидаемому'
